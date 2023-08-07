import json
import os
import re
import time

from openpyxl.reader.excel import load_workbook

from base.decorators import authenticated_core
from base.handler import BaseHandler
from base.res import res_func, res_fail_func
from base.utils import mongo_helper, now_utc, get_random, get_md5
from base.xlsx import save_to_excel
from config import settings
from core.cores.func import get_obj_info, recursion_category_delete
from core.cores.service import CoreService
from core.dicts.service import DictService
from core.menus.service import MenuService
from core.params.service import ParamService
from core.users.service import UserService


class AddHandler(BaseHandler):
    """
        添加
        post -> /core/add/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode("utf-8")
        req_data = json.loads(data)

        # 校验唯一性
        unique = True

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        add_json = {}
        # 判断哪些字段需要添加
        for item in module["table_json"]:
            # 判断是否存在UID，且不能编辑
            if item["name"] == "uid" and item["edit"] is False:
                add_json["uid"] = current_user["_id"]
            # 判断是否创建时间，且不能编辑
            if item["name"] == "add_time" and item["edit"] is False:
                add_json["add_time"] = now_utc()
            value = req_data.get(item["name"])
            if value is not None:
                # 加入字段
                add_json[item["name"]] = value

                # 判断是否用户模块
                if "User" == module["mid"] and item["name"] == "password":
                    # 加密密码
                    add_json["password"] = get_md5(add_json["password"])

                if item["type"] == 2 or item["type"] == 9:
                    # Int/Object类型转换
                    add_json[item['name']] = int(add_json.get(item['name']))
                elif item["type"] == 3:
                    # Float类型转换
                    add_json[item['name']] = float(add_json.get(item['name']))
                elif item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                    # 图片/富文本/视频 地址转换
                    url = add_json.get(item['name'])
                    add_json[item['name']] = url.replace(settings['SITE_URL'], "#URL#")
                elif item["type"] == 12 or item["type"] == 14:
                    # 多图片/多视频 地址转换
                    url_list = add_json.get(item['name'])
                    add_json[item['name']] = ",".join(url_list).replace(settings['SITE_URL'], "#URL#").split(",")
                # 判断是否有唯一字段校验
                if item["unique"]:
                    query = await mongo_helper.fetch_one(module["mid"], {item['name']: add_json.get(item['name'])})
                    if query is not None:
                        res['code'] = 10004
                        res['message'] = '已存在'
                        unique = False
                        break
            else:
                # 使用默认值
                if item.get("default") is not None and item["edit"] is True:
                    add_json[item["name"]] = item["default"]
        if unique:
            # 不用校验，直接入库
            _id = await mongo_helper.get_next_id(module["mid"])
            add_json["_id"] = _id
            await mongo_helper.insert_one(module["mid"], add_json)
            if module["cache"] == 1:
                # 删除缓存
                await CoreService.remove_category(module["mid"])
                # 判断需要删除缓存的模块
                if module["mid"] == 'Param':
                    # 删除参数缓存
                    ParamService.remove_params(None)
                elif 'DictValue' == module["mid"]:
                    dict_value = await mongo_helper.fetch_one("DictValue", {"_id": _id})
                    if dict_value is not None:
                        # 删除字典
                        DictService.delete_cache(dict_value["type_name"])
                elif module["mid"] == "Menu":
                    # 删除菜单缓存
                    MenuService.remove_menus()
        self.write(res)


class RecursionDeleteHandler(BaseHandler):
    """
        递归删除
        post -> /core/recursionDelete/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode("utf-8")
        req_data = json.loads(data)
        _id = req_data.get("id")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        if _id is not None:
            _id = int(_id)
            # 递归删除数据
            await recursion_category_delete(module["mid"], _id)
            # 删除缓存
            await CoreService.remove_category(module["mid"])
            if module["mid"] == "Menu":
                # 删除菜单缓存
                MenuService.remove_menus()
        else:
            res = res_fail_func(None)
        self.write(res)


class DeleteHandler(BaseHandler):
    """
        删除
        post -> /core/delete/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode("utf-8")
        req_data = json.loads(data)
        _id = req_data.get("id")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        if _id is not None:
            _id = int(_id)
            if module["cache"] == 1:
                # 删除缓存
                await CoreService.remove_obj(module["mid"], _id)
                if 'Param' == module["mid"]:
                    param = await mongo_helper.fetch_one("Param", {"_id": _id})
                    if param is not None:
                        # 删除参数缓存
                        ParamService.remove_params(param["key"])
                elif 'DictType' == module["mid"]:
                    dict_type = await mongo_helper.fetch_one("DictType", {"_id": _id})
                    if dict_type is not None:
                        # 删除字典
                        DictService.delete_cache(dict_type["name"])
                elif 'DictValue' == module["mid"]:
                    dict_value = await mongo_helper.fetch_one("DictValue", {"_id": _id})
                    if dict_value is not None:
                        # 删除字典
                        DictService.delete_cache(dict_value["type_name"])
                elif 'User' == module["mid"]:
                    # 删除用户缓存
                    UserService.delete_cache(_id)
            # 删除数据
            await mongo_helper.delete_one(module["mid"], {"_id": _id})
        else:
            res = res_fail_func(None)
        self.write(res)


class BatchDeleteHandler(BaseHandler):
    """
        批量删除
        post -> /core/batchDelete/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode('utf-8')
        req_data = json.loads(data)
        ids = req_data.get("ids")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        for _id in ids:
            _id = int(_id)
            # 删除数据
            await mongo_helper.delete_one(module["mid"], {"_id": _id})
            if module["cache"] == 1:
                # 删除缓存
                await CoreService.remove_obj(module["mid"], _id)
                if 'User' == module["mid"]:
                    # 删除用户缓存
                    UserService.delete_cache(_id)

        self.write(res)


class UpdateHandler(BaseHandler):
    """
        修改
        post -> /core/update/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode("utf-8")
        req_data = json.loads(data)
        _id = req_data.get("id")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        if _id is not None:
            _id = int(_id)
            update_json = {}
            # 判断哪些字典需要修改
            for item in module["table_json"]:
                value = req_data.get(item["name"])
                if value is not None:
                    # 带.字典不能修改
                    if item["name"].find(".") > -1:
                        # 处理下一个字段
                        continue
                    # 处理其他字段
                    update_json[item["name"]] = value

                    # 判断是否用户模块
                    if "User" == module["mid"] and item["name"] == "password":
                        # 加密密码
                        update_json["password"] = get_md5(update_json["password"])

                    if item["type"] == 2:
                        # Int类型转换
                        update_json[item['name']] = int(update_json[item['name']])
                    elif item["type"] == 3:
                        # Float类型转换
                        update_json[item['name']] = float(update_json[item['name']])
                    elif item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                        # 图片/富文本/视频 地址转换
                        url = update_json.get(item['name'])
                        update_json[item['name']] = url.replace(settings['SITE_URL'], "#URL#")
                    elif item["type"] == 12 or item["type"] == 14:
                        # 多图片/多视频 地址转换
                        url_list = update_json.get(item['name'])
                        update_json[item['name']] = ",".join(url_list).replace(settings['SITE_URL'], "#URL#").split(",")
            # 修改数据
            await mongo_helper.update_one(module["mid"], {"_id": _id}, {"$set": update_json})
            if module["cache"] == 1:
                # 删除缓存
                await CoreService.remove_obj(module["mid"], _id)
                if 'Param' == module["mid"]:
                    param = await mongo_helper.fetch_one("Param", {"_id": _id})
                    if param is not None:
                        # 删除参数缓存
                        ParamService.remove_params(param["key"])
                elif 'DictValue' == module["mid"]:
                    dict_value = await mongo_helper.fetch_one("DictValue", {"_id": _id})
                    if dict_value is not None:
                        # 删除字典缓存
                        DictService.delete_cache(dict_value["type_name"])
                elif module["mid"] == "Menu":
                    # 删除菜单缓存
                    MenuService.remove_menus()
                elif 'User' == module["mid"]:
                    # 删除用户缓存
                    UserService.delete_cache(_id)
        else:
            res = res_fail_func(None)
        self.write(res)


class BatchUpdateHandler(BaseHandler):
    """
        批量修改
        post -> /core/batchUpdate/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode("utf-8")
        req_data = json.loads(data)
        ids = req_data.get("ids")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        for _id in ids:
            _id = int(_id)
            update_json = {}
            # 判断哪些字典需要修改
            for item in module["table_json"]:
                value = req_data.get(item["name"])
                if value is not None:
                    update_json[item["name"]] = value
            if len(update_json) > 0:
                # 修改数据
                await mongo_helper.update_one(module["mid"], {"_id": _id}, {"$set": update_json})
                if module["cache"] == 1:
                    # 删除缓存
                    await CoreService.remove_obj(module["mid"], _id)
        self.write(res)


class ListHandler(BaseHandler):
    """
        列表
        post -> /core/list/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode('utf-8')
        req_data = json.loads(data)
        current_page = req_data.get("currentPage", 1)
        page_size = req_data.get("pageSize", 10)
        search_key = req_data.get("searchKey")
        sort_field = req_data.get("sortField", "_id")
        sort_order = req_data.get("sortOrder", "descending")
        orgs = req_data.get("orgs")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        # 需要替换地址
        replace_url = []
        # 需要批量替换地址
        replace_urls = []
        # 需要替换对象ID
        objects = []

        # 综合查询条件
        query_criteria = {"_id": {"$ne": "sequence_id"}}
        if search_key is not None:
            query_list = []
            for item in module["table_json"]:
                # 是否是查询条件
                if item["query"]:
                    if item["type"] == 1 or item["type"] == 7:
                        query_list.append({item["name"]: re.compile(search_key)})
            if len(query_list) > 0:
                query_criteria["$or"] = query_list

        # 单独查询条件
        for item in module["table_json"]:
            if item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                # 需要替换地址
                replace_url.append(item["name"])
            elif item["type"] == 12 or item["type"] == 14:
                # 需要批量替换地址
                replace_urls.append(item["name"])
            elif item["type"] == 9:
                # 对象替换成详情
                if item.get("key") is not None:
                    objects.append({"field": item["name"], "mid": item.get("key")})

            # 是否是查询条件
            if item.get("single_query"):
                value = req_data.get(item["name"])
                if value is not None:
                    if item["type"] == 2:
                        # Int类型
                        query_criteria[item["name"]] = int(value)
                    elif item["type"] == 4:
                        # 查询字典集合
                        query_criteria[item["name"]] = {"$in": value}
                    elif item["type"] == 10:
                        # 查询分类
                        value = req_data.get(item["name"])
                        if value is not None and len(value) > 0:
                            query_criteria[item["name"]] = value
                    else:
                        # 其他查询
                        query_criteria[item["name"]] = value
        # 排序条件
        if sort_field == "_id":
            sort_data = [("add_time", -1 if sort_order == 'descending' else 1),
                         ("sort", -1 if sort_order == 'descending' else 1),
                         ("_id", -1 if sort_order == 'descending' else 1)]
        else:
            sort_data = [(sort_field, -1 if sort_order == 'descending' else 1),
                         ("sort", -1 if sort_order == 'descending' else 1),
                         ("_id", -1 if sort_order == 'descending' else 1)]
        # 按部门查询
        if orgs is not None:
            query_criteria["orgs"] = {"$in": [orgs]}

        # 查询分页数据
        page_data = await mongo_helper.fetch_page_info(module["mid"], query_criteria,
                                                       sort_data, page_size,
                                                       current_page)
        # 查询总数
        total = await mongo_helper.fetch_count_info(module["mid"], query_criteria)

        results = []
        for item in page_data.get("list", []):
            item["id"] = item["_id"]
            if item.get("password") is not None:
                item['password'] = ""
            item.pop("_id")
            # 替换地址
            for url_key in replace_url:
                if item.get(url_key) is not None:
                    item[url_key] = item[url_key].replace("#URL#", settings['SITE_URL'])
            # 批量替换地址
            for url_key in replace_urls:
                if item.get(url_key) is not None:
                    item[url_key] = ",".join(item[url_key]).replace("#URL#", settings['SITE_URL']).split(",")
            # 查询对象详情
            for obj in objects:
                if item.get(obj["field"]) is not None:
                    info = await get_obj_info(obj["mid"], item[obj["field"]])
                    if info is not None:
                        # 匹配字段
                        for table in module["table_json"]:
                            if table["name"].find(obj["mid"]) > -1:
                                # 匹配上字段，给该字段赋值
                                item[table["name"]] = info[table["name"].replace(obj["mid"] + ".", "")]
            results.append(item)
        data = {
            "total": total,
            "size": page_size,
            "current": current_page,
            "results": results
        }

        res['data'] = data
        self.write(res)


class GetListHandler(BaseHandler):
    """
        所有列表
        post -> /core/getList/
    """

    @authenticated_core
    async def post(self):
        res = res_func([])
        data = self.request.body.decode('utf-8')
        req_data = json.loads(data)
        orgs = req_data.get("orgs")

        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        # 需要替换地址
        replace_url = []
        # 需要批量替换地址
        replace_urls = []
        # 需要替换对象ID
        objects = []

        # 查询条件
        query_criteria = {"_id": {"$ne": "sequence_id"}}
        # 单独查询条件
        for item in module["table_json"]:
            if item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                # 需要替换地址
                replace_url.append(item["name"])
            elif item["type"] == 12 or item["type"] == 14:
                # 需要批量替换地址
                replace_urls.append(item["name"])
            elif item["type"] == 9:
                # 对象替换成详情
                if item.get("key") is not None:
                    objects.append({"field": item["name"], "mid": item.get("key")})
            # 是否是查询条件
            if item.get("single_query"):
                value = req_data.get(item["name"])
                if value is not None:
                    if item["type"] == 2:
                        # Int类型
                        query_criteria[item["name"]] = int(value)
                    elif item["type"] == 4:
                        # 查询字典集合
                        query_criteria[item["name"]] = {"$in": value}
                    elif item["type"] == 10:
                        # 查询分类
                        value = req_data.get(item["name"])
                        if value is not None and len(value) > 0:
                            query_criteria[item["name"]] = value
                    else:
                        # 其他查询
                        query_criteria[item["name"]] = value
        # 按部门查询
        if orgs is not None:
            query_criteria["orgs"] = {"$in": [orgs]}

        query = await mongo_helper.fetch_all(module["mid"], query_criteria, [("sort", -1), ("_id", -1)])
        results = []
        for item in query:
            item["id"] = item["_id"]
            if item.get("password") is not None:
                item['password'] = ""
            item.pop("_id")
            # 替换地址
            for url_key in replace_url:
                if item.get(url_key) is not None:
                    item[url_key] = item[url_key].replace("#URL#", settings['SITE_URL'])
            # 批量替换地址
            for url_key in replace_urls:
                if item.get(url_key) is not None:
                    item[url_key] = ",".join(item[url_key]).replace("#URL#", settings['SITE_URL']).split(",")
            # 查询对象详情
            for obj in objects:
                if item.get(obj["field"]) is not None:
                    info = await get_obj_info(obj["mid"], item[obj["field"]])
                    if info is not None:
                        # 匹配字段
                        for table in module["table_json"]:
                            if table["name"].find(obj["mid"]) > -1:
                                # 匹配上字段，给该字段赋值
                                item[table["name"]] = info[table["name"].replace(obj["mid"] + ".", "")]
            results.append(item)
        res['data'] = results
        self.write(res)


class GetCategoryHandler(BaseHandler):
    """
        获取分类列表
        post -> /core/getCategory/
    """

    @authenticated_core
    async def post(self):
        res = res_func([])
        data = self.request.body.decode('utf-8')
        req_data = json.loads(data)
        # 是否查询统计
        statistics = req_data.get("statistics")
        # 是否查询分类
        type_id = req_data.get("type_id", 0)

        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        # 需要替换地址
        replace_url = []
        # 需要批量替换地址
        replace_urls = []

        # 模块字段检查
        for item in module["table_json"]:
            if item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                # 需要替换地址
                replace_url.append(item["name"])
            elif item["type"] == 12 or item["type"] == 14:
                # 需要批量替换地址
                replace_urls.append(item["name"])
        # 查询分类关联的数据
        statistics_mid = None
        if statistics is not None:
            code = await mongo_helper.fetch_one('Code',
                                                {"table_json": {"$elemMatch": {"type": 10, "key": module["mid"]}}})
            if code is not None:
                statistics_mid = code["mid"]
        # 获取分类列表
        query = await CoreService.get_category(module["mid"], int(type_id))
        results = []
        for item in query:
            # 替换地址
            for url_key in replace_url:
                if item.get(url_key) is not None:
                    item[url_key] = item[url_key].replace("#URL#", settings['SITE_URL'])
            # 批量替换地址
            for url_key in replace_urls:
                if item.get(url_key) is not None:
                    item[url_key] = ",".join(item[url_key]).replace("#URL#", settings['SITE_URL']).split(",")
            # 查询分类下使用数量
            if statistics is not None and statistics_mid is not None:
                item["Statistics.num"] = await mongo_helper.fetch_count_info(statistics_mid, {
                    "categorys": {"$elemMatch": {"value": item["id"]}}})
            results.append(item)
        # 如果是部门数据，需要根据用户权限筛选
        if module["mid"] == "Department":
            if current_user.get("orgs") is not None:
                flag = False
                orgs = []
                for item in results:
                    # 判断用户所在的部门
                    if current_user["orgs"][len(current_user["orgs"]) - 1]["value"] == item["id"]:
                        flag = True
                    if flag:
                        item["show"] = True
                        orgs.append(item)
                    else:
                        item["show"] = False
                        orgs.append(item)
                res['data'] = orgs
            else:
                res['data'] = results
        else:
            res['data'] = results
        self.write(res)


class GetInfoHandler(BaseHandler):
    """
        获取详情
        post -> /core/getInfo/
    """

    @authenticated_core
    async def post(self):
        res = res_func({})
        data = self.request.body.decode('utf-8')
        req_data = json.loads(data)
        _id = req_data.get("id")

        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        # 需要替换地址
        replace_url = []
        # 需要批量替换地址
        replace_urls = []
        # 需要替换对象ID
        objects = []

        if _id is not None:
            _id = int(_id)
            # 模块字段检查
            for item in module["table_json"]:
                if item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                    # 是否需要替换地址
                    replace_url.append(item["name"])
                elif item["type"] == 12 or item["type"] == 14:
                    # 需要批量替换地址
                    replace_urls.append(item["name"])
                elif item["type"] == 9:
                    # 对象替换成详情
                    if item.get("key") is not None:
                        objects.append({"field": item["name"], "mid": item.get("key")})
            if module["cache"] == 1:
                query = await CoreService.get_obj(module["mid"], _id)
            else:
                query = await mongo_helper.fetch_one(module["mid"], {"_id": _id})
            if query is not None:
                query["id"] = query["_id"]
                if query.get("password") is not None:
                    query['password'] = ""
                query.pop("_id")
                # 替换地址
                for url_key in replace_url:
                    if query.get(url_key) is not None:
                        query[url_key] = query[url_key].replace("#URL#", settings['SITE_URL'])
                # 批量替换地址
                for url_key in replace_urls:
                    if query.get(url_key) is not None:
                        query[url_key] = ",".join(query[url_key]).replace("#URL#", settings['SITE_URL']).split(",")
                # 查询对象详情
                for obj in objects:
                    if query.get(obj["field"]) is not None:
                        info = await get_obj_info(obj["mid"], query[obj["field"]])
                        if info is not None:
                            # 匹配字段
                            for table in module["table_json"]:
                                if table["name"].find(obj["mid"]) > -1:
                                    # 匹配上字段，给该字段赋值
                                    query[table["name"]] = info[table["name"].replace(obj["mid"] + ".", "")]
                res['data'] = query
        self.write(res)


class ImportDataHandler(BaseHandler):
    """
        导入数据
        post -> /core/importData/
    """

    @authenticated_core
    async def post(self):
        res = res_func(None)
        # 将导入的数据归类
        field_key = self.request.arguments.get("field_key")
        field_value = self.request.arguments.get("field_value")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        # 需要导入的字段
        fields = []

        # 字典查询条件
        for item in module["table_json"]:
            # 处理导入的字段
            if item["name"] in module["import_rule"]["fields"]:
                fields.append({"type": item["type"], "name": item["name"], "key": item["key"]})

        time_path = time.strftime("%Y%m%d", time.localtime())
        upload_path = os.path.join(os.path.dirname(__file__), settings['SAVE_URL'] + '/files', time_path)
        # 判断文件夹是否存在
        if os.path.isdir(upload_path) is False:
            os.makedirs(upload_path)

        file_metas = self.request.files.get('file', None)
        if not file_metas:
            res['code'] = 50000
            res['message'] = '上传文件为空'
        else:
            for meta in file_metas:
                suffix = meta['filename'].split(".")[1].lower()
                filename = get_random(6) + "." + suffix
                file_path = os.path.join(upload_path, filename)
                with open(file_path, 'wb') as up:
                    up.write(meta['body'])
                wb = load_workbook(file_path)
                # 获得sheet对象
                ws = wb.active

                # 导入规则
                rules = module["import_rule"]["rules"]
                for rule in rules:
                    # 开始读取行数
                    start_row = int(rule["start_row"])
                    end_row = int(rule["end_row"])
                    # 结束读取行数
                    max_row = ws.max_row
                    if end_row > 0:
                        max_row = end_row
                    while start_row <= max_row:
                        column_index = 1
                        add_json = {}
                        for field in fields:
                            value = ws.cell(row=start_row, column=column_index).value
                            if value is not None:
                                # 判断类型
                                if field["name"].find(".") > -1:
                                    # 对象处理
                                    id_key = None
                                    name = field["name"].replace(field["key"]+".", "")
                                    for item in module["table_json"]:
                                        if item["key"] == field["key"]:
                                            if id_key is None:
                                                id_key = item["name"]
                                        if item["name"] == field["name"]:
                                            # 判断对象ID是否已赋值
                                            if add_json.get(id_key) is None:
                                                obj = await mongo_helper.fetch_one(field["key"], {name: value})
                                                if obj is not None:
                                                    add_json[id_key] = obj["_id"]
                                            break
                                elif field['type'] == 4:
                                    # 多字典，转换
                                    values = value.split(",")
                                    items = []
                                    for text in values:
                                        dict_value = await mongo_helper.fetch_one("DictValue",
                                                                                  {"type_name": field["key"],
                                                                                   "text": text})
                                        if dict_value is not None:
                                            items.append(dict_value["value"])
                                    add_json[field['name']] = items
                                elif field['type'] == 5:
                                    # 单字典，转换
                                    dict_value = await mongo_helper.fetch_one("DictValue",
                                                                              {"type_name": field["key"],
                                                                               "text": value})
                                    if dict_value is not None:
                                        add_json[field['name']] = dict_value["value"]
                                elif field['type'] == 10:
                                    # 分类数据转换
                                    values = value.split(",")
                                    categorys = []
                                    level = 1
                                    for item in values:
                                        # 查询分类数据
                                        category = await mongo_helper.fetch_one(field["key"], {"name": item, "level": level})
                                        if category is not None:
                                            categorys.append({"text": category["name"], "value": category["_id"]})
                                        level += 1
                                    add_json[field['name']] = categorys
                                elif field['type'] == 6 or field['type'] == 8 or field['type'] == 13:
                                    # 单文件，转换
                                    add_json[field['name']] = value.replace(settings['SITE_URL'], "#URL#")
                                elif field['type'] == 12 or field['type'] == 14:
                                    # 多文件，转换
                                    values = value.split(",")
                                    urls = []
                                    for url in values:
                                        urls.append(url.replace(settings['SITE_URL'], "#URL#"))
                                    add_json[field['name']] = urls
                                elif field['type'] == 2 or field['type'] == 9 or field['type'] == 15:
                                    # 转int
                                    add_json[field['name']] = int(value)
                                elif field['type'] == 3:
                                    # 转float
                                    add_json[field['name']] = float(value)
                                else:
                                    add_json[field['name']] = str(value)
                            # 判断是否User模块
                            if module['mid'] == "User":
                                # 直接初始化密码和角色
                                add_json["password"] = get_md5("123456")
                                add_json["roles"] = ["user"]
                                add_json["status"] = "1"
                            column_index += 1
                        # 加入数据库
                        _id = await mongo_helper.get_next_id(module["mid"])
                        add_json["_id"] = _id
                        if field_key is not None and field_value is not None:
                            # 导入数据归类
                            add_json[field_key[0].decode('utf-8')] = int(field_value[0].decode('utf-8'))
                        await mongo_helper.insert_one(module["mid"], add_json)
                        start_row = start_row + 1
                res['message'] = '导入成功'
        self.write(res)


class ExportDataHandler(BaseHandler):
    """
        导出数据
        post -> /core/exportData/
    """

    @authenticated_core
    async def post(self):
        res = res_func(None)
        data = self.request.body.decode('utf-8')
        req_data = json.loads(data)
        search_key = req_data.get("searchKey")
        sort_field = req_data.get("sortField", "_id")
        sort_order = req_data.get("sortOrder", "descending")
        orgs = req_data.get("orgs")

        # 当前用户信息
        current_user = self.current_user
        # 获取模块信息
        module = current_user["module"]

        # 头部数据
        head_row = []
        # 需要导出的字段
        fields = []

        # 需要替换地址
        replace_url = []
        # 需要批量替换地址
        replace_urls = []
        # 需要替换对象ID
        objects = []

        # 综合查询条件
        query_criteria = {"_id": {"$ne": "sequence_id"}}
        if search_key is not None:
            query_list = []
            for item in module["table_json"]:
                # 是否是查询条件
                if item["query"]:
                    if item["type"] == 1 or item["type"] == 7:
                        query_list.append({item["name"]: re.compile(search_key)})
            if len(query_list) > 0:
                query_criteria["$or"] = query_list

        # 单独查询条件
        for item in module["table_json"]:
            # 处理导出的字段
            if item["name"] in module["export_rule"]["fields"]:
                head_row.append(item["remarks"])
                fields.append({"type": item["type"], "name": item["name"], "key": item["key"]})

            if item["type"] == 6 or item["type"] == 8 or item["type"] == 13:
                # 需要替换地址
                replace_url.append(item["name"])
            elif item["type"] == 12 or item["type"] == 14:
                # 需要批量替换地址
                replace_urls.append(item["name"])
            elif item["type"] == 9:
                # 对象替换成详情
                if item.get("key") is not None:
                    objects.append({"field": item["name"], "mid": item.get("key")})

            # 是否是查询条件
            if item.get("single_query"):
                value = req_data.get(item["name"])
                if value is not None:
                    if item["type"] == 2:
                        # Int类型
                        query_criteria[item["name"]] = int(value)
                    elif item["type"] == 4:
                        # 查询字典集合
                        query_criteria[item["name"]] = {"$in": value}
                    elif item["type"] == 10:
                        # 查询分类
                        value = req_data.get(item["name"])
                        if value is not None and len(value) > 0:
                            query_criteria[item["name"]] = value
                    else:
                        # 其他查询
                        query_criteria[item["name"]] = value
        # 判断是否有需要导出的字段
        if len(head_row) > 0:
            # 排序条件
            if sort_field == "_id":
                sort_data = [("add_time", -1 if sort_order == 'descending' else 1),
                             ("sort", -1 if sort_order == 'descending' else 1),
                             ("_id", -1 if sort_order == 'descending' else 1)]
            else:
                sort_data = [(sort_field, -1 if sort_order == 'descending' else 1),
                             ("sort", -1 if sort_order == 'descending' else 1),
                             ("_id", -1 if sort_order == 'descending' else 1)]
            # 按部门查询
            if orgs is not None:
                query_criteria["orgs"] = {"$in": [orgs]}

            # 查询数据
            query = await mongo_helper.fetch_all(module["mid"], query_criteria, sort_data)

            results = []
            for item in query:
                item["id"] = item["_id"]
                if item.get("password") is not None:
                    item['password'] = ""
                item.pop("_id")
                # 替换地址
                for url_key in replace_url:
                    if item.get(url_key) is not None:
                        item[url_key] = item[url_key].replace("#URL#", settings['SITE_URL'])
                # 批量替换地址
                for url_key in replace_urls:
                    if item.get(url_key) is not None:
                        item[url_key] = ",".join(item[url_key]).replace("#URL#", settings['SITE_URL']).split(",")
                # 查询对象详情
                for obj in objects:
                    if item.get(obj["field"]) is not None:
                        info = await get_obj_info(obj["mid"], item[obj["field"]])
                        if info is not None:
                            # 匹配字段
                            for table in module["table_json"]:
                                if table["name"].find(obj["mid"]) > -1:
                                    # 匹配上字段，给该字段赋值
                                    item[table["name"]] = info[table["name"].replace(obj["mid"] + ".", "")]
                results.append(item)
            # 导出的数据
            datas = []
            for item in results:
                obj = []
                for field in fields:
                    if item.get(field['name']) is not None:
                        # 判断类型，有些特殊的字段需要处理
                        if field['type'] == 4:
                            # 多字典，转换
                            texts = []
                            for value in item[field['name']]:
                                dict_value = await mongo_helper.fetch_one("DictValue",
                                                                          {"type_name": field["key"], "value": value})
                                if dict_value is not None:
                                    texts.append(dict_value["text"])
                            if len(texts) > 0:
                                obj.append(",".join(texts))
                        elif field['type'] == 5:
                            # 单字典，转换
                            dict_value = await mongo_helper.fetch_one("DictValue",
                                                                      {"type_name": field["key"],
                                                                       "value": item[field['name']]})
                            if dict_value is not None:
                                obj.append(dict_value["text"])
                        elif field['type'] == 10:
                            # 类型转换
                            texts = []
                            for category in item[field['name']]:
                                texts.append(category["text"])
                            if len(texts) > 0:
                                obj.append(",".join(texts))
                        elif field['type'] == 12:
                            # 多图片转换
                            obj.append(",".join(item[field['name']]))
                        elif field['type'] == 14:
                            # 多文件转换
                            obj.append(",".join(item[field['name']]))
                        else:
                            # 转成字符串
                            obj.append(str(item[field['name']]))
                    else:
                        obj.append("")
                datas.append(obj)
            # 保存文件位置
            save_excel_name = str(now_utc()) + ".xlsx"
            res["data"] = save_to_excel(datas, head_row, save_excel_name)
        else:
            res['code'] = 50000
            res['message'] = "导出失败"
        self.write(res)
