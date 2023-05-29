import os

from openpyxl import Workbook

from config import settings


# 保存文件到xlsx
def save_to_excel(records, head_row, save_excel_name="save.xlsx"):
    # 新建一个workbook
    wb = Workbook()
    path = os.path.join(os.path.dirname(__file__), settings['SAVE_URL'] + '/export')
    # 判断文件夹是否存在
    if os.path.isdir(path) is False:
        os.makedirs(path)
    # 设置文件输出路径与名称
    dist_filename = os.path.join(os.path.dirname(__file__), settings['SAVE_URL'] + '/export', save_excel_name)

    # 第一个sheet是ws
    ws = wb.worksheets[0]
    # 写第一行，标题行
    for h_x in range(1, len(head_row) + 1):
        ws.cell(row=1, column=h_x).value = head_row[h_x - 1]
    # 写第二行及其以后的那些行
    i = 2
    for record in records:
        record_list = record
        for x in range(1, len(record_list) + 1):
            ws.cell(row=i, column=x).value = record_list[x - 1]
        i += 1
    # 写文件
    wb.save(filename=dist_filename)
    return settings['SITE_URL'] + "/export/" + save_excel_name
